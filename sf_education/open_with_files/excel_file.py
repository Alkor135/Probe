import openpyxl as xl

file = xl.load_workbook('Sample.xlsx')
sheet = file.active
# print(sheet['A'][2].value)

# # Суммировать значения выручки за 2016 год по регионам
region_16_revenue = {}
for row in sheet.iter_rows(min_row=2):
    if row[0].value:  # Проверка на 'none' (если не 'none')
        region = row[0].value
        rev2016 = row[3].value
        default_value = 0
        region_16_revenue[region] = region_16_revenue.get(region, default_value) + rev2016

print(region_16_revenue)


regionWPerStore = {}

for row in sheet.iter_rows(min_row=2):
    if row[0].value:
        storeName = row[0].value + " " + row[1].value
        values = {'2015': 0, '2016': 0, '2017': 0}

        revenue2015 = row[2].value
        revenue2016 = row[3].value
        revenue2017 = row[4].value

        values['2015'] = 1
        values['2016'] = revenue2016/revenue2015-1
        values['2017'] = revenue2017/revenue2016-1

        regionWPerStore[storeName] = values

print(regionWPerStore)

